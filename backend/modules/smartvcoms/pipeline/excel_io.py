"""Excel input loader for VCOMS processing."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


DASHBOARD_COLUMNS = [
    "STT",
    "Mã hồ sơ",
    "Phòng",
    "CIF",
    "Tên KH",
    "Số tiền GN",
    "Luồng GN",
    "CB HTTD",
    "LĐP/KSV HTTD",
    "SLA",
    "Hồ sơ đến",
    "Tiếp nhận",
    "Phê duyệt",
    "Ký số",
    "Giải ngân",
    "Tiến độ HS",
    "Số tài khoản",
    "Trạng thái Luồng",
    "Thời gian nhận email",
    "Cập nhật cuối",
    "EntryID",
    "Thời gian SLA",
    "Thời gian SLA HTTD",
    "Thời gian SLA BGĐ",
]


@dataclass
class WorkbookInputs:
    """Container for input sheets/config used by VCOMS processor."""

    data: pd.DataFrame
    dashboard: pd.DataFrame
    config_cb: pd.DataFrame
    config_ld: pd.DataFrame
    vn: pd.DataFrame
    phongban: pd.DataFrame
    keywords: dict[str, str]
    sla_config: dict[str, float]


def _read_sheet(path: Path, sheet_name: str) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")


def _extract_keywords(vn_df: pd.DataFrame) -> dict[str, str]:
    mapping = {}
    key_col = next((c for c in vn_df.columns if "ID" in str(c)), vn_df.columns[0])
    value_col = next((c for c in vn_df.columns if "Cột B" in str(c)), vn_df.columns[1])
    for _, row in vn_df.iterrows():
        key = row.get(key_col)
        value = row.get(value_col)
        if pd.notna(key) and pd.notna(value):
            mapping[str(key).strip()] = str(value).strip()
    return mapping


def _extract_sla_config_from_sheet(path: Path) -> dict[str, float]:
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Config_CB"]
    try:
        o1 = float(ws["O1"].value) if ws["O1"].value is not None else 45.0
        o2 = float(ws["O2"].value) if ws["O2"].value is not None else 180.0
        o3 = float(ws["O3"].value) if ws["O3"].value is not None else 1.0
        o5 = float(ws["O5"].value) if ws["O5"].value is not None else 0.8
        o6 = float(ws["O6"].value) if ws["O6"].value is not None else 0.2
    finally:
        wb.close()
    return {
        "sla_default": o1,  # VBA O1
        "sla_max": o2,      # VBA O2
        "return_factor": o3,  # VBA O3
        "alloc_httd": o5,   # VBA O5
        "alloc_bgd": o6,    # VBA O6
    }


def _ensure_dashboard_columns(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()
    for col in DASHBOARD_COLUMNS:
        if col not in out.columns:
            out[col] = pd.NA
    return out[DASHBOARD_COLUMNS]


def read_workbook_inputs(path: str | Path) -> WorkbookInputs:
    """Read workbook sheets and parse processor config without mutating source file."""
    p = Path(path)
    data = _read_sheet(p, "Data")
    dashboard = _ensure_dashboard_columns(_read_sheet(p, "Dashboard"))
    config_cb = _read_sheet(p, "Config_CB")
    config_ld = _read_sheet(p, "Config_LD")
    vn = _read_sheet(p, "VN")
    phongban = _read_sheet(p, "PhongBan")

    return WorkbookInputs(
        data=data,
        dashboard=dashboard,
        config_cb=config_cb,
        config_ld=config_ld,
        vn=vn,
        phongban=phongban,
        keywords=_extract_keywords(vn),
        sla_config=_extract_sla_config_from_sheet(p),
    )
